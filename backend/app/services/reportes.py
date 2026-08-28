"""Generación de reportes de vigilancia exportables (PDF / Excel) — HU-17 / RF-13."""
from __future__ import annotations

import io
from datetime import datetime

from sqlalchemy import func
from sqlalchemy.orm import Session

from ..enums import NivelRiesgo
from ..models import Alerta, Comunidad, Medicion, Reservorio, Ubigeo


def _consolidado(db: Session, ubigeo_id: int, periodo: str) -> list[dict]:
    """Consolida por comunidad las mediciones y alertas de un periodo YYYY-MM."""
    anio, mes = (int(x) for x in periodo.split("-"))
    filas: list[dict] = []
    comunidades = (
        db.query(Comunidad).filter(Comunidad.ubigeo_id == ubigeo_id)
        .order_by(Comunidad.nombre).all()
    )
    for c in comunidades:
        res_ids = [r.reservorio_id for r in db.query(Reservorio).filter_by(comunidad_id=c.comunidad_id)]
        if not res_ids:
            continue
        q = db.query(Medicion).filter(
            Medicion.reservorio_id.in_(res_ids),
            func.extract("year", Medicion.fecha_hora) == anio,
            func.extract("month", Medicion.fecha_hora) == mes,
        )
        mediciones = q.all()
        total = len(mediciones)
        rojas = sum(1 for m in mediciones if m.nivel_riesgo == NivelRiesgo.ROJO)
        amarillas = sum(1 for m in mediciones if m.nivel_riesgo == NivelRiesgo.AMARILLO)
        verdes = total - rojas - amarillas
        alertas = (
            db.query(func.count(Alerta.alerta_id))
            .join(Medicion, Alerta.medicion_id == Medicion.medicion_id)
            .filter(Medicion.reservorio_id.in_(res_ids)).scalar()
        )
        filas.append({
            "comunidad": c.nombre, "mediciones": total,
            "verdes": verdes, "amarillas": amarillas, "rojas": rojas,
            "alertas": alertas or 0,
        })
    return filas


def generar_excel(db: Session, ubigeo_id: int, periodo: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    ubigeo = db.get(Ubigeo, ubigeo_id)
    filas = _consolidado(db, ubigeo_id, periodo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vigilancia"
    ws["A1"] = f"Yakuni — Reporte de vigilancia de la calidad del agua"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Distrito: {ubigeo.distrito} · Provincia: {ubigeo.provincia} · Periodo: {periodo}"
    ws["A3"] = f"Generado: {datetime.now():%Y-%m-%d %H:%M}"

    encabezados = ["Comunidad", "Mediciones", "Verdes", "Amarillas", "Rojas", "Alertas"]
    ws.append([])
    ws.append(encabezados)
    hdr_fill = PatternFill("solid", fgColor="0E7490")
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
    for f in filas:
        ws.append([f["comunidad"], f["mediciones"], f["verdes"], f["amarillas"], f["rojas"], f["alertas"]])
    for col in "ABCDEF":
        ws.column_dimensions[col].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_pdf(db: Session, ubigeo_id: int, periodo: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    ubigeo = db.get(Ubigeo, ubigeo_id)
    filas = _consolidado(db, ubigeo_id, periodo)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title="Reporte de vigilancia Yakuni")
    styles = getSampleStyleSheet()
    elems = [
        Paragraph("Yakuni — Reporte de vigilancia de la calidad del agua", styles["Title"]),
        Paragraph(f"Distrito: {ubigeo.distrito} · Provincia: {ubigeo.provincia}", styles["Normal"]),
        Paragraph(f"Periodo: {periodo} · Generado: {datetime.now():%Y-%m-%d %H:%M}", styles["Normal"]),
        Spacer(1, 16),
    ]
    data = [["Comunidad", "Medic.", "Verdes", "Amarillas", "Rojas", "Alertas"]]
    for f in filas:
        data.append([f["comunidad"], f["mediciones"], f["verdes"], f["amarillas"], f["rojas"], f["alertas"]])
    if len(data) == 1:
        data.append(["(sin datos en el periodo)", "", "", "", "", ""])

    tabla = Table(data, hAlign="LEFT")
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0E7490")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    elems.append(tabla)
    elems.append(Spacer(1, 16))
    elems.append(Paragraph(
        "Documento consolidado para remisión a la DIRESA/DESA. "
        "Trazabilidad detección–acción–verificación conforme al D.S. N.° 031-2010-SA.",
        styles["Italic"]))
    doc.build(elems)
    return buf.getvalue()
